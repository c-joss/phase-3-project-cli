import json
import questionary
from customer import Customer, Rate, TariffRate, Manager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime
import re
import os

DATA_FILE = "data/rates.json"
TARIFF_FILE = "data/tariff.json"
EXPORT_DIR = "exports"

EXPORT_HEADERS = [
    "POL",
    "POD",
    "Container",
    "Freight USD",
    "OTHC AUD",
    "DOC AUD",
    "CMR AUD",
    "AMS USD",
    "LSS USD",
    "DTHC",
    "Free Time",
]

EXPORT_HEADERS_WITH_CUSTOMER = [
    "Customer",
] + EXPORT_HEADERS


def load_data():
    try:
        with open(DATA_FILE, "r") as f:
            data = json.load(f)
            customers = []
            for c in data:
                customer = Customer(c["name"])
                for r in c["rates"]:
                    rate = Rate(
                        r["load_port"],
                        r["destination_port"],
                        r.get("container_type", ""),
                        r["freight_usd"],
                        r["othc_aud"],
                        r["doc_aud"],
                        r["cmr_aud"],
                        r["ams_usd"],
                        r["lss_usd"],
                        r["dthc"],
                        r["free_time"],
                    )
                    customer.add_rate(rate)
                customers.append(customer)
            return customers
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def save_data(customers):
    with open(DATA_FILE, "w") as f:
        json.dump([c.to_dict() for c in customers], f, indent=4)


def load_tariff():
    try:
        with open(TARIFF_FILE, "r") as f:
            data = json.load(f)
            tariffs = []
            for r in data:
                rate = TariffRate(
                    r["load_port"],
                    r["destination_port"],
                    r["container_type"],
                    {
                        "freight_usd": r["freight_usd"],
                        "othc_aud": r["othc_aud"],
                        "doc_aud": r["doc_aud"],
                        "cmr_aud": r["cmr_aud"],
                        "ams_usd": r["ams_usd"],
                        "lss_usd": r["lss_usd"],
                        "dthc": r["dthc"],
                        "free_time": r["free_time"],
                    },
                )
                tariffs.append(rate)
            return tariffs
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def save_tariff(tariffs):
    with open(TARIFF_FILE, "w") as f:
        json.dump([rate.to_dict() for rate in tariffs], f, indent=4)


def format_rate_choice(rate, index):
    return f"{index + 1}: {rate.load_port} to {rate.destination_port} ({rate.container_type})"


class TariffManager(Manager):
    def __init__(self):
        super().__init__()
        (
            self.valid_load_ports,
            self.valid_dest_ports,
            self.valid_containers,
            self.valid_dthc,
        ) = load_constants()
        self.load_tariffs()

    def load_tariffs(self):
        self.items = load_tariff()

    def save_tariffs(self):
        save_tariff(self.items)

    def add_tariffs(self, load_port, destination_port, container_type, values):
        rate = TariffRate(load_port, destination_port, container_type, values)
        self.add(rate)
        self.save_tariffs()

    def view_tariffs(self):
        if not self.items:
            print("\n No Tariff rate found.")
            return

        for r in self.items:
            print(str(r))

    def delete_tariff(self):
        if not self.items:
            print("\n No Tariff rates to delete.")
            return
        choices = [format_rate_choice(r, idx) for idx, r in enumerate(self.items)]
        selected = questionary.select("Select Tariff to delete:", choices=choices).ask()
        rate_idx = int(selected.split(":")[0]) - 1
        confirm = questionary.confirm("Confirm to delete tariff?").ask()
        if confirm:
            deleted = self.items.pop(rate_idx)
            self.save_tariffs()
            print(
                f"\n Deleted tariffs: {deleted.load_port} to {deleted.destination_port} ({deleted.container_type})\n"
            )
        else:
            print("\nCancelled.\n")

    def export_tariff_rates(self):
        export_rates_to_excel(self.items, filename_prefix="Tariff_Rates")

    def import_tariff_rates(self):
        file_path = questionary.text(
            "Enter path to Excel file to import:", default=f"{EXPORT_DIR}/"
        ).ask()

        try:
            wb = load_workbook(filename=file_path)
        except Exception as e:
            print(f"\n Could not open file: {e}\n")
            return

        ws = wb.active

        import_count = 0

        for row in ws.iter_rows(min_row=4, values_only=True):
            (
                load_port,
                destination_port,
                container_type,
                freight_usd,
                othc_aud,
                doc_aud,
                cmr_aud,
                ams_usd,
                lss_usd,
                dthc,
                free_time,
            ) = row

            load_port = load_port.upper()
            destination_port = destination_port.upper()

            existing_rate = next(
                (
                    r
                    for r in self.items
                    if r.load_port == load_port
                    and r.destination_port == destination_port
                    and r.container_type == container_type
                ),
                None,
            )

            if existing_rate:
                if (
                    existing_rate.to_dict()
                    == TariffRate(
                        load_port,
                        destination_port,
                        container_type,
                        {
                            "freight_usd": freight_usd,
                            "othc_aud": othc_aud,
                            "doc_aud": doc_aud,
                            "cmr_aud": cmr_aud,
                            "ams_usd": ams_usd,
                            "lss_usd": lss_usd,
                            "dthc": dthc,
                            "free_time": free_time,
                        },
                    ).to_dict()
                ):
                    print(
                        f"Tariff for {load_port} → {destination_port} ({container_type}) already exists. Skipping."
                    )
                    continue

                confirm_replace = questionary.confirm(
                    f"Tariff for {load_port} → {destination_port} ({container_type}) exists with different rates. Replace?"
                ).ask()

                if confirm_replace:
                    self.items = [
                        r
                        for r in self.items
                        if not (
                            r.load_port == load_port
                            and r.destination_port == destination_port
                            and r.container_type == container_type
                        )
                    ]
                    print(
                        f"Replaced tariff for {load_port} → {destination_port} ({container_type})"
                    )
                else:
                    print(
                        f"Skipped updating {load_port} → {destination_port} ({container_type})"
                    )
                    continue

            if load_port not in self.valid_load_ports:
                confirm_add = questionary.confirm(
                    f"\nLoad Port '{load_port}' not in valid list. Would you like to add it?"
                ).ask()
                if confirm_add:
                    self.valid_load_ports.append(load_port)
                    save_constants(
                        self.valid_load_ports,
                        self.valid_dest_ports,
                        self.valid_containers,
                        self.valid_dthc,
                    )
                    print(f"Added load port: {load_port}")
                else:
                    print(f"Rate for '{load_port}' not added.")
                    continue

            if destination_port not in self.valid_dest_ports:
                confirm_add = questionary.confirm(
                    f"\nDestination Port '{destination_port}' not in valid list. Would you like to add it?"
                ).ask()
                if confirm_add:
                    self.valid_dest_ports.append(destination_port)
                    save_constants(
                        self.valid_load_ports,
                        self.valid_dest_ports,
                        self.valid_containers,
                        self.valid_dthc,
                    )
                    print(f"Added destination port: {destination_port}")
                else:
                    print(f"Rate for '{destination_port}' not added.")
                    continue

            rate = TariffRate(
                load_port,
                destination_port,
                container_type,
                {
                    "freight_usd": freight_usd,
                    "othc_aud": othc_aud,
                    "doc_aud": doc_aud,
                    "cmr_aud": cmr_aud,
                    "ams_usd": ams_usd,
                    "lss_usd": lss_usd,
                    "dthc": dthc,
                    "free_time": free_time,
                },
            )

            self.add(rate)
            import_count += 1

        if import_count > 0:
            self.save_tariffs()

        print(f"\n Import complete: {import_count} new tariff rates imported.\n")

    def get_valid_ports(self):
        return (
            self.valid_load_ports,
            self.valid_dest_ports,
            self.valid_containers,
            self.valid_dthc,
        )


def export_rates_to_excel(rates, filename_prefix):

    if not rates:
        print("\n No rates found.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"

    ws["A1"] = f"{filename_prefix} Export"
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])

    ws.append(EXPORT_HEADERS)

    for cell in ws[3]:
        cell.font = Font(bold=True)

    for rate in rates:
        ws.append(rate.to_row())

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        col_letter = column_cells[0].column_letter
        ws.column_dimensions[col_letter].width = length + 2

    current_date = datetime.now().strftime("%d_%m_%Y")
    filename = f"{EXPORT_DIR}/{filename_prefix}_{current_date}.xlsx"
    wb.save(filename)

    print(f"\n Exported to {len(rates)} rates to {filename}\n")


def rate_values_prompt(load_ports, dest_ports, containers, dthc_values, defaults=None):
    defaults = defaults or {}

    load_port = questionary.autocomplete(
        "Enter Load Port:",
        choices=load_ports,
        default=defaults.get("load_port", ""),
        validate=lambda text: text in load_ports or "Please select a valid port",
    ).ask()

    destination_port = questionary.autocomplete(
        "Enter Destination Port:",
        choices=dest_ports,
        default=defaults.get("destination_port", ""),
        validate=lambda text: text in dest_ports or "Please select a valid port",
    ).ask()

    container_type_default = defaults.get("container_type", "")
    container_type = questionary.select(
        "Select Container Type:",
        choices=containers,
        default=(
            container_type_default if container_type_default in containers else None
        ),
    ).ask()

    freight_usd = questionary.text(
        "Freight (USD):", default=str(defaults.get("freight_usd", ""))
    ).ask()
    othc_aud = questionary.text(
        "OTHC (AUD):", default=str(defaults.get("othc_aud", ""))
    ).ask()
    doc_aud = questionary.text(
        "DOC (AUD):", default=str(defaults.get("doc_aud", ""))
    ).ask()
    cmr_aud = questionary.text(
        "CMR (AUD):", default=str(defaults.get("cmr_aud", ""))
    ).ask()
    ams_usd = questionary.text(
        "AMS (USD):", default=str(defaults.get("ams_usd", ""))
    ).ask()
    lss_usd = questionary.text(
        "LSS (USD):", default=str(defaults.get("lss_usd", ""))
    ).ask()
    dthc_default = defaults.get("dthc", "")
    dthc = questionary.select(
        "Select DTHC Terms:",
        choices=dthc_values,
        default=dthc_default if dthc_default in dthc_values else None,
    ).ask()
    free_time = questionary.text(
        "Free Time:", default=str(defaults.get("free_time", ""))
    ).ask()

    return {
        "load_port": load_port,
        "destination_port": destination_port,
        "container_type": container_type,
        "freight_usd": freight_usd,
        "othc_aud": othc_aud,
        "doc_aud": doc_aud,
        "cmr_aud": cmr_aud,
        "ams_usd": ams_usd,
        "lss_usd": lss_usd,
        "dthc": dthc,
        "free_time": free_time,
    }


def load_constants():
    try:
        with open("data/data_constants.json", "r") as f:
            data = json.load(f)
            return (
                data["VALID_LOAD_PORTS"],
                data["VALID_DEST_PORTS"],
                data["VALID_CONTAINERS"],
                data["VALID_DTHC"],
            )
    except (FileNotFoundError, json.JSONDecodeError):
        return [], [], [], []


def save_constants(load_ports, dest_ports, containers, dthc_values):
    with open("data/data_constants.json", "w") as f:
        json.dump(
            {
                "VALID_LOAD_PORTS": load_ports,
                "VALID_DEST_PORTS": dest_ports,
                "VALID_CONTAINERS": containers,
                "VALID_DTHC": dthc_values,
            },
            f,
            indent=4,
        )
