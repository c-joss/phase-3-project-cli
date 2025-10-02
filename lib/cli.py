import questionary
from lib.db.models import Session, Customer, Rate
from tabulate import tabulate
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime
import re
from lib.helpers import (
    load_data, get_valid_ports, rate_values_prompt, format_rate_choice,
    TariffManager, export_rates_to_excel, export_tariff_rates_to_excel
)


def main_menu():
    while True:
        choice = questionary.select(
            "What would you like to do?",
            choices=[
                "Add Rate",
                "View Rates",
                "Edit Rates",
                "Delete Rate",
                "Export Quote to Excel",
                "Export Customers by Destination Port",
                "Import Quote from Excel",
                "Manage Tariff Rates",
                "Exit",
            ],
        ).ask()

        if choice == "Add Rate":
            add_rate()
        elif choice == "View Rates":
            view_rates()
        elif choice == "Edit Rates":
            edit_rates()
        elif choice == "Delete Rate":
            delete_rate()
        elif choice == "Export Quote to Excel":
            export_quote()
        elif choice == "Export Customers by Destination Port":
            export_by_destination()
        elif choice == "Import Quote from Excel":
            import_quote()
        elif choice == "Manage Tariff Rates":
            manage_tariff_rate()
        elif choice == "Exit":
            print("Exiting Rate Manager App")
            break


def add_rate():
    customer_name = questionary.text("Enter customer name:").ask().strip().upper()
    load_ports, dest_ports, containers, dthc_values = get_valid_ports()
    values = rate_values_prompt(load_ports, dest_ports, containers, dthc_values)

    s = Session()
    try:
        customer = s.query(Customer).filter_by(name=customer_name).first()
        if not customer:
            customer = Customer(name=customer_name)
            s.add(customer)
            s.flush()

        existing = s.query(Rate).filter_by(
            customer_id=customer.id,
            load_port=values["load_port"],
            destination_port=values["destination_port"],
            container_type=values["container_type"],
        ).first()

        if existing:
            print("\n A rate for this route and container type already exist.")
            print("Existing:", format_rate_choice(existing, 0))
            if questionary.confirm("Do you want to update the existing rate?").ask():
                for k in ("freight_usd","othc_aud","doc_aud","cmr_aud","ams_usd","lss_usd","dthc","free_time"):
                    setattr(existing, k, values[k])
            else:
                print("\n Skipped.\n")
                return
        else:
            s.add(Rate(customer_id=customer.id, **values))

        s.commit()
        print("\n Rate saved.\n")
    finally:
        s.close()

def view_rates():
    customers = load_data()
    if not customers:
        print("\n No rates found. Please add rates first")
        return

    headers = ["POL","POD","Container","Freight USD","OTHC AUD","DOC AUD","CMR AUD","AMS USD","LSS USD","DTHC","Free Time"]
    for customer in customers:
        print(f"\nCustomer: {customer.name}")
        rows = [[
            r.load_port, r.destination_port, r.container_type,
            r.freight_usd, r.othc_aud, r.doc_aud, r.cmr_aud,
            r.ams_usd, r.lss_usd, r.dthc, r.free_time
        ] for r in customer.rates]
        print(tabulate(rows, headers=headers, tablefmt="grid"))

def edit_rates():
    customers = load_data()
    if not customers:
        print("\n No rates found.")
        return

    customer_choices = [c.name for c in customers]
    customer_name = questionary.autocomplete(
        "Select Customer:",
        choices=customer_choices,
        validate=lambda text: text in customer_choices
        or "Please select a valid customer",
    ).ask()
    customer = next(c for c in customers if c.name == customer_name)

    if not customer.rates:
        print("\n Customer has no rates.")
        return

    rate_choices = [
        f"{idx + 1}: {r.load_port} to {r.destination_port} ({r.container_type})"
        for idx, r in enumerate(customer.rates)
    ]
    selected = questionary.select("Select Rate to Edit:", choices=rate_choices).ask()
    rate_idx = int(selected.split(":")[0]) - 1
    rate = customer.rates[rate_idx]

    load_ports, dest_ports, containers, dthc_values = get_valid_ports()
    values = rate_values_prompt(
        load_ports, dest_ports, containers, dthc_values,
        defaults={
            "load_port": rate.load_port, "destination_port": rate.destination_port,
            "container_type": rate.container_type, "freight_usd": rate.freight_usd,
            "othc_aud": rate.othc_aud, "doc_aud": rate.doc_aud, "cmr_aud": rate.cmr_aud,
            "ams_usd": rate.ams_usd, "lss_usd": rate.lss_usd, "dthc": rate.dthc,
            "free_time": rate.free_time,
        },
    )

    s = Session()
    try:
        db_rate = s.query(Rate).filter_by(
            customer_id=rate.customer_id,
            load_port=rate.load_port,
            destination_port=rate.destination_port,
            container_type=rate.container_type,
        ).first()
        for k in ("load_port","destination_port","container_type","freight_usd","othc_aud",
                  "doc_aud","cmr_aud","ams_usd","lss_usd","dthc","free_time"):
            setattr(db_rate, k, values[k])
        s.commit()
        print("\n Rate updated.\n")
    finally:
        s.close()

def delete_rate():
    customers = load_data()
    if not customers:
        print("\n No rates found.")
        return

    customer_choices = [c.name for c in customers]
    customer_name = questionary.select(
        "Select Customer:", choices=customer_choices
    ).ask()
    customer = next(c for c in customers if c.name == customer_name)

    if not customer.rates:
        print("\n Customer has no rates.")
        return
    rate_choices = [
        f"{idx + 1}: {r.load_port} to {r.destination_port} ({r.container_type})"
        for idx, r in enumerate(customer.rates)
    ]
    selected = questionary.select("Select Rate to Delete:", choices=rate_choices).ask()
    rate_idx = int(selected.split(":")[0]) - 1
    target = customer.rates[rate_idx]    

    if not questionary.confirm("Confirm to delete rate?").ask():
        print("\nCancelled. \n")
        return

    s = Session()
    try:
        obj = s.query(Rate).filter_by(
            customer_id=target.customer_id,
            load_port=target.load_port,
            destination_port=target.destination_port,
            container_type=target.container_type,
        ).first()
        if obj:
            s.delete(obj)
            s.commit()
            print(f"\n Deleted rate: {target.load_port} to {target.destination_port}\n")
        else:
            print("\n Rate not found.\n")
    finally:
        s.close()


def export_quote():
    customers = load_data()

    if not customers:
        print("\n No rates found.")
        return

    customer_choices = [c.name for c in customers]
    customer_name = questionary.select(
        "Select Customer:", choices=customer_choices
    ).ask()

    customer = next(c for c in customers if c.name == customer_name)

    if not customer.rates:
        print("\n Customer has no rates.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"

    ws["A1"] = f"Customer: {customer.name}"
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])

    ws.append(EXPORT_HEADERS)

    for cell in ws[3]:
        cell.font = Font(bold=True)

    for r in customer.rates:
        row = [r.load_port, r.destination_port, r.container_type,
        r.freight_usd, r.othc_aud, r.doc_aud, r.cmr_aud,
        r.ams_usd, r.lss_usd, r.dthc, r.free_time]
        ws.append(row)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        col_letter = column_cells[0].column_letter
        ws.column_dimensions[col_letter].width = length + 2

    safe_name = re.sub(r"\W+", "_", customer_name)
    current_date = datetime.now().strftime("%d_%m_%Y")
    filename = f"{EXPORT_DIR}/Quote_{safe_name}_{current_date}.xlsx"
    wb.save(filename)

    print(f"\n Quote exported to {filename}\n")


def export_by_destination():
    customers = load_data()

    if not customers:
        print("\n No rates found.")
        return

    all_dest_ports = sorted(set(r.destination_port for c in customers for r in c.rates))

    dest_port = questionary.select(
        "Select Destination Port to export:", choices=all_dest_ports
    ).ask()

    wb = Workbook()
    ws = wb.active
    ws.title = "Quote"

    ws["A1"] = f"Destination Port: {dest_port}"
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([])

    ws.append(EXPORT_HEADERS_WITH_CUSTOMER)

    for cell in ws[3]:
        cell.font = Font(bold=True)

    rates_exported = 0

    for customer in customers:
        matching_rates = [r for r in customer.rates if r.destination_port == dest_port]
        for rate in matching_rates:
            ws.append([customer.name] + rate.to_row())
            rates_exported += 1

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        col_letter = column_cells[0].column_letter
        ws.column_dimensions[col_letter].width = length + 2

    if rates_exported == 0:
        print(f"\n No rates found for {dest_port}.")
        return

    safe_dest = re.sub(r"\W+", "_", dest_port)
    current_date = datetime.now().strftime("%d_%m_%Y")

    filename = f"{EXPORT_DIR}/Rates_{safe_dest}_{current_date}.xlsx"
    wb.save(filename)

    print(f"\n Exported {rates_exported} rates for {dest_port} to {filename}\n")


def import_quote():
    customers = load_data()
    file_path = questionary.text(
        "Enter path to Excel file to import:", default=f"{EXPORT_DIR}/"
    ).ask()

    try:
        wb = load_workbook(filename=file_path)
    except Exception as e:
        print(f"\n Could not open file: {e}\n")
        return

    ws = wb.active

    first_header = ws["A3"].value
    is_multi_customer = (first_header or "").lower() == "customer"

    if is_multi_customer:
        print("\n Import by destination port")
    else:
        customer_choices = [c.name for c in customers]
        customer_name = questionary.select(
            "Select Customer to import rates to", choices=customer_choices
        ).ask()

        customer = next(c for c in customers if c.name == customer_name)

    import_count = 0
    skip_count = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        if is_multi_customer:
            (
                customer_name,
                load_port,
                destination_port,
                container_type,
                freight_usd,
                othc_aud,
                doc_aud,
                cmr_aud,
                ams_usd,
                lss_usd,
                dthc,
                free_time,
            ) = row

            customer_name_upper = customer_name.upper()

            existing_customer = next(
                (c for c in customers if c.name == customer_name_upper), None
            )
            if not existing_customer:
                existing_customer = Customer(customer_name_upper)
                customers.append(existing_customer)
            target_customer = existing_customer
        else:
            (
                load_port,
                destination_port,
                container_type,
                freight_usd,
                othc_aud,
                doc_aud,
                cmr_aud,
                ams_usd,
                lss_usd,
                dthc,
                free_time,
            ) = row
            target_customer = customer

        new_rate = Rate(
            load_port,
            destination_port,
            container_type,
            freight_usd,
            othc_aud,
            doc_aud,
            cmr_aud,
            ams_usd,
            lss_usd,
            dthc,
            free_time,
        )

        existing_rate = next(
            (
                r
                for r in target_customer.rates
                if r.load_port == load_port
                and r.destination_port == destination_port
                and r.container_type == container_type
            ),
            None,
        )

        if existing_rate:
            if existing_rate.to_dict() == new_rate.to_dict():
                print(
                    f"Rate for {load_port} to {destination_port} ({container_type}) already exists - no changes were made."
                )
                skip_count += 1
                continue
            else:
                print(
                    f"\n Rate exists for {load_port} to {destination_port} ({container_type})"
                )
                print("Existing:".ljust(12), existing_rate)
                print("New:".ljust(12), new_rate)
                confirm = questionary.confirm("Replace existing rate?").ask()
                if confirm:
                    target_customer.rates = [
                        r
                        for r in target_customer.rates
                        if not (
                            r.load_port == load_port
                            and r.destination_port == destination_port
                            and r.container_type == container_type
                        )
                    ]
                    target_customer.add_rate(new_rate)
                    print("Rate replaced.")
                    import_count += 1
                else:
                    print("Rate skipped")
                    skip_count += 1
        else:
            target_customer.add_rate(new_rate)
            print(
                f"Imported new rate: {load_port} to {destination_port} ({container_type})"
            )
            import_count += 1

    save_data(customers)

    print(f"\n Import complete: {import_count} new/replaced, {skip_count} skipped.\n")


def manage_tariff_rate():
    tariff_manager = TariffManager()

    while True:
        action = questionary.select(
            "What would you like to do?",
            choices=[
                "View Tariff Rates",
                "Add Tariff Rate",
                "Delete Tariff Rate",
                "Export Tariff Rates to Excel",
                "Import Tariff Rates from Excel",
                "Back to Main Menu",
            ],
        ).ask()

        if action == "View Tariff Rates":
            tariff_manager.view_tariffs()

        elif action == "Add Tariff Rate":
            load_ports, dest_ports, containers, dthc_values = (
                tariff_manager.get_valid_ports()
            )
            values = rate_values_prompt(load_ports, dest_ports, containers, dthc_values)

            tariff_manager.add_tariffs(
                values["load_port"],
                values["destination_port"],
                values["container_type"],
                {
                    "freight_usd": values["freight_usd"],
                    "othc_aud": values["othc_aud"],
                    "doc_aud": values["doc_aud"],
                    "cmr_aud": values["cmr_aud"],
                    "ams_usd": values["ams_usd"],
                    "lss_usd": values["lss_usd"],
                    "dthc": values["dthc"],
                    "free_time": values["free_time"],
                },
            )
            print("\nTariff Added.\n")
        elif action == "Delete Tariff Rate":
            tariff_manager.delete_tariff()
        elif action == "Export Tariff Rates to Excel":
            tariff_manager.export_tariff_rates()
        elif action == "Import Tariff Rates from Excel":
            tariff_manager.import_tariff_rates()
        elif action == "Back to Main Menu":
            break


if __name__ == "__main__":
    main_menu()
