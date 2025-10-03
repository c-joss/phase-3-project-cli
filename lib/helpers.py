from __future__ import annotations
from pathlib import Path
from typing import List, Tuple, Dict, Any, Optional
import json
import questionary
from openpyxl import Workbook
from lib.db.models import Session, Customer, Rate, Tariff
from sqlalchemy.orm import joinedload
from openpyxl import load_workbook
from datetime import datetime

DATA_CONSTANTS = Path(__file__).resolve().parents[1] / "data" / "data_constants.json"

def _load_constants() -> Dict[str, Any]:
    if DATA_CONSTANTS.exists():
        try:
            return json.loads(DATA_CONSTANTS.read_text())
        except Exception:
            pass
    return {
        "VALID_LOAD_PORTS": ["MELBOURNE", "SYDNEY", "BRISBANE"],
        "VALID_DEST_PORTS": ["TAICHUNG", "SHANGHAI", "NINGBO", "SHEKOU", "TOKYO"],
        "VALID_CONTAINERS": ["20GP", "40GP", "40HC", "20RE", "40REHC"],
        "VALID_DTHC": ["COLLECT", "PREPAID"],
    }

def get_valid_ports() -> Tuple[List[str], List[str], List[str], List[str]]:
    const = _load_constants()
    return (
        const.get("VALID_LOAD_PORTS", []),
        const.get("VALID_DEST_PORTS", []),
        const.get("VALID_CONTAINERS", []),
        const.get("VALID_DTHC", []),
    )

def load_data() -> List[Customer]:
    s = Session()
    try:
        return (
            s.query(Customer)
             .options(joinedload(Customer.rates))
             .order_by(Customer.name)
             .all()
        )
    finally:
        s.close()

def save_data(_customers: Any) -> None:
    return

def _ask_choice(prompt: str, choices: List[str]) -> str:
    return questionary.select(prompt, choices=choices).ask()

def _ask_text(prompt: str, default: Optional[str] = None) -> str:
    return questionary.text(prompt, default=default or "").ask() or ""

def _ask_confirm(prompt: str, default: bool = False) -> bool:
    return questionary.confirm(prompt, default=default).ask()

def rate_values_prompt(
    load_ports: List[str],
    dest_ports: List[str],
    containers: List[str],
    dthc_values: List[str],
    defaults: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    defaults = defaults or {}
    load_port = _ask_choice("Load Port:", load_ports)
    dest_port = _ask_choice("Destination Port:", dest_ports)
    container = _ask_choice("Container Type:", containers)

    def _num(name: str, default: Any = "") -> float:
        val = _ask_text(f"{name}:", str(defaults.get(name, default)))
        try:
            return float(val)
        except Exception:
            return 0.0

    freight_usd = _num("freight_usd", defaults.get("freight_usd", "0"))
    othc_aud = _num("othc_aud", defaults.get("othc_aud", "0"))
    doc_aud = _num("doc_aud", defaults.get("doc_aud", "0"))
    cmr_aud = _num("cmr_aud", defaults.get("cmr_aud", "0"))
    ams_usd = _num("ams_usd", defaults.get("ams_usd", "0"))
    lss_usd = _num("lss_usd", defaults.get("lss_usd", "0"))
    dthc = _ask_choice("DTHC:", dthc_values)
    free_time = _ask_text("Free Time (e.g., 14 Days):", defaults.get("free_time", "14 Days"))

    return {
        "load_port": load_port,
        "destination_port": dest_port,
        "container_type": container,
        "freight_usd": freight_usd,
        "othc_aud": othc_aud,
        "doc_aud": doc_aud,
        "cmr_aud": cmr_aud,
        "ams_usd": ams_usd,
        "lss_usd": lss_usd,
        "dthc": dthc,
        "free_time": free_time,
    }

def format_rate_choice(r: Rate, idx: int) -> str:
    return f"{idx+1}: {r.load_port} → {r.destination_port} ({r.container_type})  USD {r.freight_usd:.2f}"

class TariffManager:
    def __init__(self) -> None:
        self.items: List[Tariff] = []

    def load_tariffs(self) -> None:
        s = Session()
        try:
            self.items = s.query(Tariff).all()
        finally:
            s.close()

    def save_tariffs(self) -> None:
        return

    def add_tariffs(
        self,
        load_port: str,
        destination_port: str,
        container_type: str,
        values: Dict[str, Any],
    ) -> None:
        s = Session()
        try:
            t = Tariff(
                load_port=load_port,
                destination_port=destination_port,
                container_type=container_type,
                freight_usd=float(values["freight_usd"]),
                othc_aud=float(values["othc_aud"]),
                doc_aud=float(values["doc_aud"]),
                cmr_aud=float(values["cmr_aud"]),
                ams_usd=float(values["ams_usd"]),
                lss_usd=float(values["lss_usd"]),
                dthc=str(values["dthc"]).upper(),
                free_time=str(values["free_time"]),
            )
            s.add(t)
            s.commit()
        finally:
            s.close()
        self.load_tariffs()

    def delete_tariff(self, selected_index: int) -> bool:
        if not self.items:
            print("\n No Tariff rates to delete.")
            return False
        i = max(0, min(selected_index, len(self.items) - 1))
        to_delete = self.items[i]
        s = Session()
        try:
            obj = s.get(Tariff, to_delete.id)
            if obj:
                s.delete(obj)
                s.commit()
                print(
                    f"\n Deleted tariff: {to_delete.load_port} → "
                    f"{to_delete.destination_port} ({to_delete.container_type})\n"
                )
                return True
            return False
        finally:
            s.close()
            self.load_tariffs()

    def import_tariff_rates(self):
        file_path = questionary.text(
            "Enter path to Tariff Excel file:", default=f"{EXPORTS_DIR}/"
        ).ask()

        try:
            wb = load_workbook(filename=file_path)
        except Exception as e:
            print(f"\n Could not open file: {e}\n")
            return

        ws = wb.active
        new_count = updated_count = skipped_count = 0
        s = Session()
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None or all(v is None for v in row):
                    continue

                (load_port, destination_port, container_type,
                freight_usd, othc_aud, doc_aud, cmr_aud,
                ams_usd, lss_usd, dthc, free_time) = row

                def _f(x):
                    try:
                        return float(x)
                    except Exception:
                        return 0.0

                values = dict(
                    load_port=(load_port or "").strip(),
                    destination_port=(destination_port or "").strip(),
                    container_type=(container_type or "").strip(),
                    freight_usd=_f(freight_usd),
                    othc_aud=_f(othc_aud),
                    doc_aud=_f(doc_aud),
                    cmr_aud=_f(cmr_aud),
                    ams_usd=_f(ams_usd),
                    lss_usd=_f(lss_usd),
                    dthc=str(dthc or "").upper(),
                    free_time=str(free_time or ""),
                )

                existing = s.query(Tariff).filter_by(
                    load_port=values["load_port"],
                    destination_port=values["destination_port"],
                    container_type=values["container_type"],
                ).first()

                if existing:
                    changed = any(getattr(existing, k) != v for k, v in values.items())
                    if changed:
                        for k, v in values.items():
                            setattr(existing, k, v)
                        updated_count += 1
                    else:
                        skipped_count += 1
                else:
                    s.add(Tariff(**values))
                    new_count += 1

            s.commit()
            print(f"\n Tariff import complete: {new_count} new, {updated_count} updated, {skipped_count} skipped.\n")
        finally:
            s.close()
            self.load_tariffs()

EXPORTS_DIR = Path(__file__).resolve().parents[1] / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)

def _rate_to_row(r: Rate) -> List[Any]:
    return [
        r.load_port, r.destination_port, r.container_type,
        r.freight_usd, r.othc_aud, r.doc_aud, r.cmr_aud,
        r.ams_usd, r.lss_usd, r.dthc, r.free_time
    ]

def _tariff_to_row(t: Tariff) -> List[Any]:
    return [
        t.load_port, t.destination_port, t.container_type,
        t.freight_usd, t.othc_aud, t.doc_aud, t.cmr_aud,
        t.ams_usd, t.lss_usd, t.dthc, t.free_time
    ]

def export_rates_to_excel(rates, file_prefix, directory=None) -> Path:
    outdir = Path(directory) if directory else EXPORTS_DIR
    outdir.mkdir(parents=True, exist_ok=True)

    current_date = datetime.now().strftime("%d_%m_%Y")

    out_path = outdir / f"{file_prefix}_{current_date}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"
    headers = [
        "Load Port", "Destination Port", "Container",
        "Freight USD", "OTHC AUD", "DOC AUD", "CMR AUD",
        "AMS USD", "LSS USD", "DTHC", "Free Time"
    ]
    ws.append(headers)

    for r in rates:
        if hasattr(r, "to_row") and callable(getattr(r, "to_row")):
            ws.append(list(r.to_row()))
        else:
            ws.append(_rate_to_row(r))

    wb.save(out_path)
    return out_path

def export_tariff_rates_to_excel(tariffs: List[Tariff], filename: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tariff Rates"
    headers = [
        "Load Port", "Destination Port", "Container",
        "Freight USD", "OTHC AUD", "DOC AUD", "CMR AUD",
        "AMS USD", "LSS USD", "DTHC", "Free Time"
    ]
    ws.append(headers)
    for t in tariffs:
        ws.append(_tariff_to_row(t))
    out = EXPORTS_DIR / f"{filename}.xlsx"
    wb.save(out)
    return out

def replace_or_add_rate(customer_name, values):

    name = (customer_name or "").strip().upper()
    s = Session()
    try:
        customer = s.query(Customer).filter_by(name=name).first()
        if not customer:
            customer = Customer(name=name)
            s.add(customer)
            s.flush()

        lp  = (values.get("load_port") or "").strip()
        dp  = (values.get("destination_port") or "").strip()
        ctn = (values.get("container_type") or "").strip()

        rate = s.query(Rate).filter_by(
            customer_id=customer.id,
            load_port=lp,
            destination_port=dp,
            container_type=ctn,
        ).first()

        def _f(x):
            try: return float(x)
            except Exception: return 0.0

        fields = dict(
            load_port=lp,
            destination_port=dp,
            container_type=ctn,
            freight_usd=_f(values.get("freight_usd")),
            othc_aud=_f(values.get("othc_aud")),
            doc_aud=_f(values.get("doc_aud")),
            cmr_aud=_f(values.get("cmr_aud")),
            ams_usd=_f(values.get("ams_usd")),
            lss_usd=_f(values.get("lss_usd")),
            dthc=str(values.get("dthc") or "").upper(),
            free_time=str(values.get("free_time") or ""),
            customer_id=customer.id,
        )

        if rate:
            for k, v in fields.items():
                setattr(rate, k, v)
        else:
            rate = Rate(**fields)
            s.add(rate)

        s.commit()
        s.refresh(rate)
        return rate
    finally:
        s.close()

def replace_or_add_rate(customer, new_rate, replace_existing=None):

    match = None
    for r in getattr(customer, "rates", []):
        if (
            r.load_port == new_rate.load_port and
            r.destination_port == new_rate.destination_port and
            r.container_type == new_rate.container_type
        ):
            match = r
            break

    if match:
        if replace_existing is None:
            ans = questionary.confirm(
                "A rate for this route and container exists. Replace it?"
            ).ask()
            if not ans:
                return match
        elif replace_existing is False:
            return match

        for attr in [
            "freight_usd","othc_aud","doc_aud","cmr_aud",
            "ams_usd","lss_usd","dthc","free_time"
        ]:
            setattr(match, attr, getattr(new_rate, attr))
        return match
    else:
        if not hasattr(customer, "rates"):
            customer.rates = []
        customer.rates.append(new_rate)
        return new_rate

